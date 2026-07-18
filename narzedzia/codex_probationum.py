#!/usr/bin/env python3
"""
📜 CODEX PROBATIONUM — żywy rejestr testów Imperium w Excelu

Generuje wielo-arkuszowy plik .xlsx z DWÓCH źródeł prawdy (nigdy z pamięci):
  1. ŻYWY KOD (rejestry neuronów/strategii/Namiestnika) → arkusze-referencje,
  2. WERSJONOWANY JSONL `bibliotheca_ulpia/dane/rejestr_testow.jsonl` → arkusze wyników.
Excel = generowany WIDOK; regeneracja jedną komendą, więc nie może się rozjechać (Prawo XXI).

Uruchom:
    python narzedzia/codex_probationum.py                 # -> raporty/CODEX_PROBATIONUM.xlsx
    python narzedzia/codex_probationum.py --wyjscie X.xlsx

Architektura (Prawo I — miękki fallback):
    zbierz_arkusze()  — RDZEŃ, czysty, testowalny BEZ openpyxl (zwraca {arkusz: wiersze}).
    zapisz_xlsx()     — LENIWY import openpyxl; brak biblioteki → czytelny komunikat, nie crash.

Aby DOPISAĆ wynik testu: dopisz linię JSON do rejestr_testow.jsonl i uruchom generator.
"""
import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Windows: konsola cp1250 wywala się na emoji → wymuś UTF-8 (no-op na Linux/macOS).
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")   # type: ignore[union-attr]
    except Exception:  # noqa: BLE001
        pass

LEDGER = ROOT / "bibliotheca_ulpia" / "dane" / "rejestr_testow.jsonl"
DOMYSLNE_WYJSCIE = ROOT / "raporty" / "CODEX_PROBATIONUM.xlsx"
MATRYCA = ROOT / "docs" / "MATRYCA_KORELACJI.md"

# Legenda KATEGORIA — KOMPLETNA (wszystkie 15 liter żywych w kodzie + planowane E/G).
# Jedyne źródło prawdy: komentarz w imperium/legiony/mikro_neuron.py (utrzymuj zgodność).
KAT_NAZWA = {
    "M": "Momentum", "T": "Trend", "V": "Zmienność", "F": "Flow/Wolumen",
    "O": "On-chain", "L": "Leverage/Dźwignia", "R": "Reżim/Sentyment",
    "S": "Struktura (SMC)", "A": "Anty-manipulacja", "K": "Makro/Intermarket",
    "E": "Entropia/AI (planowane)", "G": "Geo/Regionalne (planowane)",
    "C": "Cross-sectional/Przekrój", "D": "Geometria ścieżki (Path Signature)",
    "H": "Hurst/Pamięć długiego zasięgu", "N": "Entropia/Informacja (PermEn)",
    "Z": "Zagrożenie (VPIN)",
}

# Interwały reprezentatywne dla arkusza „Interwały→Styl" (kolejność od najszybszego).
INTERWALY_REPR = ["M1", "M5", "M15", "M30", "1H", "4H", "1D", "1W"]

# Katalogi danych: nazwa_interwalu → (podkatalog, sufiks_pliku).
DANE_INTERWALY = {
    "1m": ("dane/minutowe", "_minute"),
    "1h": ("dane/godzinowe", "_1h"),
    "4h": ("dane/4h", "_4h"),
    "1d": ("dane/dzienne", "_d"),
}


# ── Pomocnicze odczyty źródeł ────────────────────────────────────────────────

def _wczytaj_ledger(sciezka: Path = LEDGER) -> list:
    """Czyta rejestr_testow.jsonl (jeden rekord na linię). Brak pliku → []."""
    if not Path(sciezka).exists():
        return []
    rekordy = []
    for linia in Path(sciezka).read_text(encoding="utf-8", errors="replace").splitlines():
        linia = linia.strip()
        if not linia:
            continue
        try:
            rekordy.append(json.loads(linia))
        except json.JSONDecodeError:
            continue   # uszkodzona linia nie zabija generatora (Prawo I)
    return rekordy


def _zywotnosc_adapterow() -> dict:
    """{neuron: nota} z ostatniego cenzusu (arena XV_ZYWY). Brak bazy → {}."""
    try:
        from imperium.biblioteki.arena_baza import pytaj_pomiary
        wiersze = pytaj_pomiary(rodzaj="XV_ZYWY", limit=200)
    except Exception:  # noqa: BLE001 — brak bazy/modułu nie może wywrócić generatora
        return {}
    out = {}
    for w in wiersze:                      # najnowsze pierwsze → pierwszy wygrywa
        out.setdefault(w["neuron"], w.get("nota", ""))
    return out


def _cele_adapterow() -> list:
    """Lista 22 kluczy-adapterów z cenzusu (fallback: znana lista, jeśli import padnie)."""
    try:
        from narzedzia.cenzus_adapterow import CELE
        return list(CELE)
    except Exception:  # noqa: BLE001
        return ["AUG-01", "C-01", "NEWS-01", "NEWS-02", "NEWS-03", "NEWS-04",
                "OC-06", "OC-07", "OC-08", "PSY-01", "PSY-02", "PSY-03", "PSY-04",
                "RADAR-01", "RADAR-02", "RADAR-03", "RADAR-04", "RADAR-05",
                "V-03", "X-28", "Z-06", "Z-07"]


def _pokrycie_danych() -> dict:
    """{(para, interwal): (najnowszy, najstarszy, liczba_barow)} z lekkiego odczytu CSV."""
    out = {}
    for nazwa_int, (podkat, suf) in DANE_INTERWALY.items():
        d = ROOT / podkat
        if not d.is_dir():
            continue
        for plik in sorted(d.glob(f"Binance_*{suf}.csv")):
            para = plik.name[len("Binance_"):-len(suf + ".csv")]
            try:
                with open(plik, encoding="utf-8", errors="replace") as fh:
                    fh.readline()                       # linia URL
                    fh.readline()                       # nagłówek kolumn
                    pierwsza = fh.readline()            # najnowszy bar (pliki malejące)
                    if not pierwsza.strip():
                        continue
                    liczba, ostatnia = 1, pierwsza
                    for linia in fh:
                        if linia.strip():
                            liczba += 1
                            ostatnia = linia
                naj = pierwsza.split(",")[1].strip()
                naj_st = ostatnia.split(",")[1].strip()
                out[(para, nazwa_int)] = (naj, naj_st, liczba)
            except (OSError, IndexError):
                continue
    return out


def _parsuj_korelacje() -> tuple:
    """Zwraca (meta, pary) z MATRYCA_KORELACJI.md — zmierzone pary nadmiarowe.

    Parsuje wiersze tabeli markdown zawierające 'A ~ B' i wartość korelacji.
    meta = jednolinijkowy opis zakresu pomiaru (data/para/interwał).
    """
    if not MATRYCA.exists():
        return ("(brak docs/MATRYCA_KORELACJI.md)", [])
    txt = MATRYCA.read_text(encoding="utf-8", errors="replace")
    meta = ""
    m = re.search(r"PIERWSZY POMIAR RZECZYWISTY[^\n]*", txt)
    if m:
        meta = m.group(0).lstrip("# ").strip()
    pary = []
    # Wiersz tabeli: | **A ~ B** | wskazniki | +0.972 | interpretacja |
    wzor = re.compile(r"\|\s*\*{0,2}([A-Z][\w-]+)\s*~\s*([A-Z][\w-]+)\*{0,2}\s*\|"
                      r"[^|]*\|\s*\*{0,2}([+-]?\d\.\d+)\*{0,2}\s*\|")
    for linia in txt.splitlines():
        mm = wzor.search(linia)
        if mm:
            pary.append((mm.group(1), mm.group(2), float(mm.group(3))))
    return (meta or "(pomiar bez opisu)", pary)


def _styl_neuronu(klucz: str, mapa: dict, wszystkie: list) -> str:
    """Czytelny styl gry neuronu z NEURONY_STYLU (fallback: wszystkie style)."""
    style = mapa.get(klucz, wszystkie)
    kolejnosc = {"SCALP": 0, "SWING": 1, "INVEST": 2}
    return "/".join(sorted(set(style), key=lambda s: kolejnosc.get(s, 9)))


def _opis_klasy(obj) -> str:
    """Pierwsza linia docstringu klasy neuronu (opis dla człowieka)."""
    doc = (type(obj).__doc__ or "").strip()
    return doc.split("\n")[0].strip() if doc else ""


# ── RDZEŃ: budowa wierszy arkuszy (bez openpyxl — testowalne) ────────────────

def zbierz_arkusze() -> dict:
    """Zwraca {nazwa_arkusza: [wiersze]}; wiersz[0] = nagłówek. Czyste dane, bez I/O Excela."""
    from imperium.legiony.rejestr import (
        wszystkie_neurony, wszyscy_zwiadowcy, raport_potencjalu, raport_elity,
        raport_profili, NEURONY_STYLU)
    try:
        from imperium.legiony.rejestr import WSZYSTKIE_STYLE
    except Exception:  # noqa: BLE001
        WSZYSTKIE_STYLE = ["SCALP", "SWING", "INVEST"]
    from imperium.legiony.strategie.rejestr_strategii import wszystkie_strategie
    from imperium.koloseum.namiestnik import profil_stylu, styl_interwalu

    neurony = wszystkie_neurony()
    zwiadowcy = wszyscy_zwiadowcy()
    strategie = wszystkie_strategie()
    pot = raport_potencjalu()
    rap_elity = raport_elity()                       # jeden odczyt, reużyty niżej (zwiadowcy)
    elity = {e["klucz"] for e in rap_elity["neurony_elite"]}
    profile = raport_profili()
    ledger = _wczytaj_ledger()
    zywotnosc = _zywotnosc_adapterow()
    cele = set(_cele_adapterow())
    pokrycie = _pokrycie_danych()
    kor_meta, kor_pary = _parsuj_korelacje()

    # indeks wyników per neuron (najświeższy A/B i IC)
    ab_per, ic_per = {}, {}
    for r in ledger:
        n = r.get("neuron", "")
        if r.get("typ") == "AB":
            ab_per.setdefault(n, []).append(r)
        elif r.get("typ") == "IC":
            ic_per.setdefault(n, []).append(r)

    def _fmt_delta(v):
        # ledger jest ręcznie dopisywalny → rekord bez delta_pp (albo null) NIE może
        # wywrócić generatora przez format '+' na nie-liczbie (recenzja 2026-07-18).
        return f"{v:+g}" if isinstance(v, (int, float)) else "?"

    def _skrot_ab(klucz):
        rs = ab_per.get(klucz, [])
        return "; ".join(f"{r.get('interwal', '?')}:{_fmt_delta(r.get('delta_pp'))}pp" for r in rs) if rs else ""

    def _skrot_ic(klucz):
        rs = ic_per.get(klucz, [])
        return "; ".join(f"{r.get('horyzont', '?')}:{r.get('ic', '?')}" for r in rs) if rs else ""

    arkusze = {}

    # ── 1. README / Legenda ──────────────────────────────────────────────────
    dzis = date.today().isoformat()
    readme = [["CODEX PROBATIONUM — żywy rejestr testów Imperium", ""]]
    readme += [
        ["Wygenerowano", dzis],
        ["Neurony (żywe)", f"{pot['neurony_lacznie']} (aktywne {pot['neurony_aktywne']}, wyciszone {pot['neurony_wyciszone']})"],
        ["Zwiadowcy", str(pot["zwiadowcy_exp"])],
        ["Strategie", str(len(strategie))],
        ["Elitarne (Prawo XX)", str(len(elity))],
        ["Profile stylu", f"SCALP {profile['scalp']} | SWING {profile['swing']} | INVEST {profile['invest']}"],
        ["", ""],
        ["ŹRÓDŁA PRAWDY", ""],
        ["Arkusze-referencje", "generowane z żywego kodu (rejestr neuronów/strategii/Namiestnik)"],
        ["Arkusze wyników", "z bibliotheca_ulpia/dane/rejestr_testow.jsonl (wersjonowany w git)"],
        ["Jak dodać wynik", "dopisz linię JSON do rejestr_testow.jsonl → uruchom generator ponownie"],
        ["", ""],
        ["LEGENDA KATEGORII", ""],
    ]
    for lit in sorted(KAT_NAZWA):
        readme.append([lit, KAT_NAZWA[lit]])
    readme += [
        ["", ""],
        ["LEGENDA STATUSÓW", ""],
        ["✅ / POMAGA / ŻYWY", "zmierzone i działa / wpięte na realnych danych"],
        ["⚪ / NEUTRALNE / CICHY", "brak wyraźnego efektu / warunkowo bramkowane"],
        ["❌ / SZKODZI", "zmierzone, pogarsza — zostaw OFF"],
        ["— / niezmierzony", "jeszcze nie zbadane (patrz arkusz Backlog)"],
    ]
    arkusze["README"] = readme

    # ── 2. Neurony (87) ──────────────────────────────────────────────────────
    naglowek_n = ["KLUCZ", "Klasa", "WSKAZNIK", "KAT", "Kategoria (nazwa)", "WAGA",
                  "LEGION", "MECHANIZM", "Dostępny", "Powód niedostępności",
                  "Elitarny", "Styl gry", "Adapter?", "Żywotność (cenzus)",
                  "A/B (ΔROI)", "IC", "Opis"]
    wiersze_n = [naglowek_n]
    for n in sorted(neurony, key=lambda x: x.KLUCZ):
        wiersze_n.append([
            n.KLUCZ, type(n).__name__, n.WSKAZNIK, n.KATEGORIA,
            KAT_NAZWA.get(n.KATEGORIA, "?"), n.WAGA, n.LEGION,
            getattr(n, "MECHANIZM", ""), "TAK" if n.DOSTEPNY else "NIE",
            getattr(n, "POWOD_NIEDOSTEPNOSCI", ""),
            "TAK" if n.KLUCZ in elity else "",
            _styl_neuronu(n.KLUCZ, NEURONY_STYLU, WSZYSTKIE_STYLE),
            "TAK" if n.KLUCZ in cele else "",
            zywotnosc.get(n.KLUCZ, ""),
            _skrot_ab(n.KLUCZ), _skrot_ic(n.KLUCZ), _opis_klasy(n),
        ])
    arkusze["Neurony"] = wiersze_n

    # ── 3. Zwiadowcy (15) ────────────────────────────────────────────────────
    wiersze_z = [["KLUCZ", "Klasa", "WSKAZNIK", "KAT", "Dostępny", "Elitarny", "Opis"]]
    elity_z = {e["klucz"] for e in rap_elity["zwiadowcy_elite"]}
    for z in sorted(zwiadowcy, key=lambda x: getattr(x, "KLUCZ", "")):
        wiersze_z.append([
            getattr(z, "KLUCZ", ""), type(z).__name__, getattr(z, "WSKAZNIK", ""),
            getattr(z, "KATEGORIA", "?"), "TAK" if getattr(z, "DOSTEPNY", True) else "NIE",
            "TAK" if getattr(z, "KLUCZ", "") in elity_z else "", _opis_klasy(z),
        ])
    arkusze["Zwiadowcy"] = wiersze_z

    # ── 4. Strategie (20) ────────────────────────────────────────────────────
    wiersze_s = [["ID", "Nazwa", "Legion", "Styl", "Interwały", "Reżim",
                  "Dźwignia", "RR", "Status", "Źródło", "#Wejście", "#Filtr",
                  "#Wyjście", "Neurony wejścia", "Neurony filtra", "Neurony wyjścia", "Warunki"]]
    for s in strategie:
        wiersze_s.append([
            s.id, s.nazwa, s.legion, s.styl, "/".join(s.interwaly),
            s.rezim_preferowany, s.dzwignia, s.rr, s.status, s.zrodlo,
            len(s.neurony_wejscie), len(s.neurony_filtr), len(s.neurony_wyjscie),
            ", ".join(s.neurony_wejscie), ", ".join(s.neurony_filtr),
            ", ".join(s.neurony_wyjscie), s.warunki,
        ])
    arkusze["Strategie"] = wiersze_s

    # ── 5. Neurony×Strategie (macierz ról) ───────────────────────────────────
    role = {}   # (klucz_neuronu, id_strategii) -> "W"/"F"/"X" (może łączyć)
    for s in strategie:
        for k in s.neurony_wejscie:
            role.setdefault((k, s.id), set()).add("W")
        for k in s.neurony_filtr:
            role.setdefault((k, s.id), set()).add("F")
        for k in s.neurony_wyjscie:
            role.setdefault((k, s.id), set()).add("X")
    uzyte = sorted({k for (k, _sid) in role})     # tylko neurony użyte w strategiach
    naglowek_m = ["KLUCZ / Strategia"] + [s.id for s in strategie]
    wiersze_m = [naglowek_m]
    for k in uzyte:
        wiersz = [k]
        for s in strategie:
            r = role.get((k, s.id))
            wiersz.append("".join(sorted(r)) if r else "")
        wiersze_m.append(wiersz)
    wiersze_m.append([""] + [""] * len(strategie))
    wiersze_m.append(["LEGENDA: W=Wejście  F=Filtr  X=Wyjście"] + [""] * len(strategie))
    arkusze["Neurony x Strategie"] = wiersze_m

    # ── 6. Adaptery / Żywotność (22) ─────────────────────────────────────────
    kat_neuronu = {n.KLUCZ: n.KATEGORIA for n in neurony}
    wiersze_a = [["KLUCZ", "KAT", "Kategoria", "Status", "Nota (ostatni cenzus)"]]
    for k in sorted(cele):
        nota = zywotnosc.get(k, "")
        status = "ŻYWY" if nota else "niezmierzony (uruchom cenzus_adapterow.py --arena)"
        wiersze_a.append([k, kat_neuronu.get(k, "?"),
                          KAT_NAZWA.get(kat_neuronu.get(k, ""), ""), status, nota])
    arkusze["Adaptery"] = wiersze_a

    # ── 7. Waluty × Interwały (pokrycie danych) ──────────────────────────────
    pary = sorted({p for (p, _i) in pokrycie})
    naglowek_w = ["Para \\ Interwał"]
    for nazwa_int in DANE_INTERWALY:
        naglowek_w += [f"{nazwa_int}: bary", f"{nazwa_int}: od", f"{nazwa_int}: do"]
    wiersze_w = [naglowek_w]
    for para in pary:
        wiersz = [para]
        for nazwa_int in DANE_INTERWALY:
            dane = pokrycie.get((para, nazwa_int))
            if dane:
                naj, naj_st, liczba = dane
                wiersz += [liczba, naj_st, naj]
            else:
                wiersz += ["—", "—", "—"]
        wiersze_w.append(wiersz)
    arkusze["Waluty x Interwaly"] = wiersze_w

    # ── 8. Interwały → Styl (Namiestnik) ─────────────────────────────────────
    wiersze_i = [["Interwał", "Styl", "Lewar cap", "Rynek", "Mnożnik progu", "Opis"]]
    for interw in INTERWALY_REPR:
        p = profil_stylu(interw)
        wiersze_i.append([interw, styl_interwalu(interw), p.lewar_cap,
                          p.rynek, p.mnoznik_progu, p.opis])
    arkusze["Interwaly -> Styl"] = wiersze_i

    # ── 9. Wyniki A/B ────────────────────────────────────────────────────────
    wiersze_ab = [["Sygnał", "Neuron", "Interwał", "Okno (barów)", "ROI B %",
                   "ROI A %", "Δ ROI (pp)", "Δ maxDD (pp)", "Werdykt", "Data", "Źródło", "Uwaga"]]
    for r in [x for x in ledger if x.get("typ") == "AB"]:
        wiersze_ab.append([
            r.get("sygnal", ""), r.get("neuron", ""), r.get("interwal", ""),
            r.get("okno_barow", ""), r.get("roi_b", ""), r.get("roi_a", ""),
            r.get("delta_pp", ""), r.get("maxdd_delta", ""), r.get("werdykt", ""),
            r.get("data", ""), r.get("zrodlo", ""), r.get("uwaga", ""),
        ])
    arkusze["Wyniki A-B"] = wiersze_ab

    # ── 10. Wyniki IC ────────────────────────────────────────────────────────
    wiersze_ic = [["Sygnał", "Neuron", "Horyzont", "IC", "Tryb", "Próg",
                   "Werdykt", "Kierunek", "Data", "Źródło", "Uwaga"]]
    for r in [x for x in ledger if x.get("typ") == "IC"]:
        wiersze_ic.append([
            r.get("sygnal", ""), r.get("neuron", ""), r.get("horyzont", ""),
            r.get("ic", ""), r.get("tryb", ""), r.get("prog", ""),
            r.get("werdykt", ""), r.get("kierunek", ""), r.get("data", ""),
            r.get("zrodlo", ""), r.get("uwaga", ""),
        ])
    arkusze["Wyniki IC"] = wiersze_ic

    # ── 11. Korelacje (z pomiaru MATRYCA_KORELACJI) ──────────────────────────
    wiersze_k = [["POMIAR", kor_meta],
                 ["", ""],
                 ["Neuron A", "Neuron B", "Korelacja", "Klasa"]]
    for a, b, c in kor_pary:
        klasa = "REDUNDANTNA (|r|>0.80)" if abs(c) > 0.80 else \
                ("DYWERSYFIKUJĄCA (|r|<0.20)" if abs(c) < 0.20 else "pośrednia")
        wiersze_k.append([a, b, c, klasa])
    wiersze_k += [["", ""],
                  ["Świeży przelicz całego roju (87) = flaga --korelacje (blokowana backtestem O(n²) — Prawo XV)", ""]]
    arkusze["Korelacje"] = wiersze_k

    # ── 12. Backlog / Planowane ──────────────────────────────────────────────
    zbadane_ab = set(ab_per) | set(ic_per)
    wiersze_b = [["Element", "Typ testu", "Status", "Priorytet", "Uwaga"]]
    # neurony-adaptery jeszcze bez A/B/IC
    for k in sorted(cele - zbadane_ab):
        wiersze_b.append([k, "A/B + IC", "PLANOWANE", "P2",
                          "adapter żywy w cenzusie, brak walidacji PnL/IC"])
    # znane długi kierunkowe
    wiersze_b += [
        ["Backtest O(n²)", "naprawa silnika", "PLANOWANE", "P1",
         "wskaźniki przeliczane per bar; blokuje długie okna 1H/4H i WFO oraz --korelacje"],
        ["Łańcuch SHA-256", "naprawa + walidacja", "PLANOWANE", "P2",
         "hash_ok=True na sztywno (dyrygent); Hermes na ścieżce decyzyjnej, ZASADA WPIĘCIA"],
        ["Włączenie flag Tier-1", "decyzja Cezara", "OCZEKUJE", "P2",
         "DVOL/STABLE pomagają na 4H; flagi opt-in OFF do decyzji"],
        ["A/B na pełnym oknie 1H/4H", "walidacja", "ZABLOKOWANE", "P1",
         "po naprawie O(n²) — dziś okno 800 barów (1H=33 dni, nierozstrzygające)"],
        # Alarmy hooka startowego = zadania, nie tapeta (ZASADA CENSORA, 2026-07-18)
        ["Konsolidacja LEKCJI pamięci (39k zn. > limit 24k)", "higiena pamięci", "ZROBIONE 2026-07-19", "—",
         "archiwizacja wg wartości retencji: 180→94 aktywnych, 86 do PAMIEC_SESJI_ARCHIWUM.md; sekcja 39244→21962 zn; nic nie skasowane (konsoliduj_lekcje, +4 testy)"],
        ["Przegląd 10 sprzeczności Refleksji W9", "higiena pamięci", "ZROBIONE 2026-07-18", "—",
         "10/10 = FP detektora (plan liczony jak negacja + stoplista bez 'jako'); naprawa u źródła, 0 sprzecznych, +3 testy granic"],
        ["Wiszący pomysł W9: Wektory semantyczne RAG lokalnie (22d)", "decyzja Cezara", "OCZEKUJE", "P3",
         "zrób / odrzuć / odłóż świadomie — FTS działa (29,7k frag.), wektory = koszt+zależności"],
    ]
    arkusze["Backlog"] = wiersze_b

    # ── 13. Sugestie / Rozbudowa (KANDYDACI — Prawo I: nie wpinamy bez weryfikacji) ──
    # Czytane z ledgera (typ=SUGESTIA) — źródło prawdy jak reszta wyników. Każda sugestia
    # rozbudowy CODEX (nowy arkusz/dział/kolumna) ląduje tu jako KANDYDAT do oceny zgodności
    # z Imperium (ZASADA CODEX PROBATIONUM w CLAUDE.md).
    wiersze_sug = [["Element / Dział", "Typ", "Uzasadnienie", "Zgodność z Imperium",
                    "Status", "Data", "Źródło"]]
    for r in [x for x in ledger if x.get("typ") == "SUGESTIA"]:
        wiersze_sug.append([
            r.get("element", ""), r.get("dzial", ""), r.get("uzasadnienie", ""),
            r.get("zgodnosc_imperium", ""), r.get("status", ""), r.get("data", ""),
            r.get("zrodlo", ""),
        ])
    arkusze["Sugestie"] = wiersze_sug

    return arkusze


# ── ZAPIS: openpyxl (leniwy import, Prawo I: fallback bez biblioteki) ─────────

def zapisz_xlsx(arkusze: dict, sciezka) -> None:
    """Zapisuje arkusze do .xlsx ze stylizacją. Brak openpyxl → czytelny wyjątek."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "Brak openpyxl — zainstaluj: pip install openpyxl "
            "(rdzeń zbierz_arkusze() działa bez niego, ale zapis .xlsx wymaga biblioteki)."
        ) from e

    sciezka = Path(sciezka)
    sciezka.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    naglowek_fill = PatternFill("solid", fgColor="1F3864")
    naglowek_font = Font(bold=True, color="FFFFFF")
    zielony = PatternFill("solid", fgColor="C6EFCE")
    szary = PatternFill("solid", fgColor="E7E6E6")
    czerwony = PatternFill("solid", fgColor="FFC7CE")

    def _kolor(tekst: str):
        t = str(tekst).upper()
        if any(s in t for s in ("POMAGA", "ŻYWY", "SKILL", "✅", "TAK")):
            return zielony
        if any(s in t for s in ("SZKODZI", "❌")):
            return czerwony
        if any(s in t for s in ("NEUTRALNE", "CICHY", "NIEZMIERZONY", "PLANOWANE",
                                "ZABLOKOWANE", "OCZEKUJE")):
            return szary
        return None

    for nazwa, wiersze in arkusze.items():
        ws = wb.create_sheet(nazwa[:31])          # limit Excela: 31 znaków
        for wiersz in wiersze:
            ws.append(wiersz)
        # nagłówek (pierwszy wiersz): pogrubienie + tło + zamrożenie
        for cell in ws[1]:
            cell.fill = naglowek_fill
            cell.font = naglowek_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        # szerokości kolumn (na podstawie treści, cap 60)
        for col in range(1, ws.max_column + 1):
            litera = get_column_letter(col)
            maks = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    maks = max(maks, len(str(v)))
            ws.column_dimensions[litera].width = min(max(maks + 2, 8), 60)
        # kolorowanie statusów (kolumny z werdyktami/statusami)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                kol = _kolor(v)
                if kol is not None:
                    ws.cell(row=row, column=col).fill = kol

    wb.save(str(sciezka))


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Generator CODEX_PROBATIONUM.xlsx (rejestr testów Imperium)")
    ap.add_argument("--wyjscie", default=str(DOMYSLNE_WYJSCIE),
                    help=f"ścieżka pliku .xlsx (domyślnie {DOMYSLNE_WYJSCIE})")
    ap.add_argument("--korelacje", action="store_true",
                    help="[PLAN] świeży przelicz korelacji całego roju — blokowany O(n²) (Prawo XV)")
    args = ap.parse_args(argv)

    print("📜 CODEX PROBATIONUM — zbieram żywe rejestry + ledger...", flush=True)
    arkusze = zbierz_arkusze()
    liczby = {k: max(len(v) - 1, 0) for k, v in arkusze.items()}
    print(f"   arkusze: {liczby}", flush=True)
    if args.korelacje:
        print("   ⚠️ --korelacje: świeży przelicz zablokowany backtestem O(n²) (Prawo XV) — używam pomiaru z MATRYCA.", flush=True)
    try:
        zapisz_xlsx(arkusze, args.wyjscie)
    except ModuleNotFoundError as e:
        print(f"❌ {e}")
        return 1
    print(f"✅ Zapisano {len(arkusze)} arkuszy → {args.wyjscie}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
